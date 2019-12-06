# coding:utf-8
from openpyxl import load_workbook
import string
import time

time_name = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(time.time()))


class Excel_r_w:

    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.wb = load_workbook(excel_path)
        self.ws = self.wb.active
        Capital = string.ascii_uppercase  # 大写字母
        self.alphabet_list = []  # 字母列表
        for i in "ABCD":
            for u in Capital:
                self.alphabet_list.append(i + u)

    def get_sheet_name(self, sheet_name):
        """
        获取 sheet_name
        :param sheet_name: str
        :return:
        """
        self.wb.get_sheet_by_name(sheet_name)

    def get_max_row(self):
        """
        返回最大的行（AG）
        :return:
        """
        return self.ws.max_row

    def get_max_col(self):
        """
        获取最大的列数（数字）
        :return:
        """
        sol_no = self.ws.max_column
        # return sol_no
        return self.alphabet_list[sol_no - 1]

    def get_min_row(self):
        """
        返回最小的行数
        :return:
        """
        return self.ws.min_row

    def get_min_col(self):
        """
        返回最小的列数
        :return:
        """
        sol_no = self.ws.min_column
        return self.alphabet_list[sol_no - 1]

    def get_value(self, index):
        """
        获取知道行数的值例如：A1
        :param index:
        :return:
        """
        return self.ws[index].value

    def write_content(self, index, content):
        """
        通过坐标写入内容，不自动保存
        例如：write_content(A1,"test") 在 A 列 1 行写入 test
        :param index: 坐标
        :param content: str
        :return:
        """
        self.ws[index] = content

    def write_content_save(self, index, content):
        """
        通过坐标写入内容，自动保存
        例如：write_content_save(A1,"test") 在 A 列 1 行写入 test 然后保存
        :param index: 坐标 str
        :param content: str
        :return:
        """
        self.ws[index] = content
        self.wb.save(self.excel_path)

    def write_datetime(self, index, content=time_name):
        """
        写入现在的时间
        :param index: 坐标 str
        :param content: 如果需要换时间格式，可以重新赋值
        :return:
        """
        self.ws[index] = content

    def save_content(self):
        """
        保存写入的内容（配合 write_content 一起使用）
        :return:
        """
        self.wb.save(self.excel_path)


if __name__ == "__main__":
    pass
